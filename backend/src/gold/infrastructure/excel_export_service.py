"""Servicio de Exportación Ejecutiva a Excel Multi-Pestaña (CU-14).

Genera un libro de trabajo Excel .xlsx estilizado profesionalmente con pestañas:
1. Resumen Ejecutivo (KPIs y Estado de Cuadre)
2. Balance por Libro (Ledger)
3. Balance por Cuenta PyG
"""

import io, logging
from pathlib import Path
import duckdb
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from src.gold.infrastructure.gold_datamart_engine import GoldDatamartEngine

logger = logging.getLogger(__name__)


def _safe(path: Path) -> str:
    return str(path.resolve()).replace("'", "''")


class ExcelExportService:
    """Generador de Informes Ejecutivos en Excel (.xlsx)."""

    def __init__(self, conn: duckdb.DuckDBPyConnection):
        self._conn = conn

    def export_gold_report(self, silver_path: Path, gold_dir: Path) -> bytes:
        """CU-14: Genera archivo Excel de informe financiero en memoria."""
        wb = openpyxl.Workbook()
        engine = GoldDatamartEngine(self._conn)

        # 1. Sheet Resumen Ejecutivo
        ws_resumen = wb.active
        ws_resumen.title = "Resumen Ejecutivo"
        self._build_resumen_sheet(ws_resumen, engine.compute_integrity_summary(silver_path))

        # 2. Sheet Balance por Libro
        ledger_path = gold_dir / "gold_balance_by_ledger.parquet"
        if ledger_path.exists():
            ws_ledger = wb.create_sheet(title="Balance por Libro")
            self._build_table_sheet(ws_ledger, ledger_path)

        # 3. Sheet Balance por Cuenta
        acct_path = gold_dir / "gold_balance_by_account.parquet"
        if acct_path.exists():
            ws_acct = wb.create_sheet(title="Balance por Cuenta")
            self._build_table_sheet(ws_acct, acct_path)

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    def _build_resumen_sheet(self, ws, integrity):
        ws.views.sheetView[0].showGridLines = True
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 25

        title_font = Font(name="Calibri", size=14, bold=True, color="1E3A8A")
        ws.cell(row=1, column=1, value="AUDITORÍA FINANCIERA - CAPA ORO").font = title_font

        headers = ["Métrica / Indicador", "Valor / Resultado"]
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1E3A8A")

        rows = [
            ("Total Cargos (Débitos)", integrity.total_debit),
            ("Total Abonos (Créditos)", integrity.total_credit),
            ("Diferencia Global", integrity.global_imbalance),
            ("Estado de Cuadre Global", "CUADRADO ✅" if integrity.is_globally_balanced else "DESCUADRADO ⚠️"),
            ("Asientos Descuadrados (#)", integrity.imbalanced_entries_count),
            ("Monto Descuadrado Total ($)", integrity.imbalanced_entries_amount),
            ("Total Registro Asientos", integrity.total_journals_count),
        ]

        for r_idx, (lbl, val) in enumerate(rows, 4):
            ws.cell(row=r_idx, column=1, value=lbl)
            c2 = ws.cell(row=r_idx, column=2, value=val)
            if isinstance(val, (int, float)):
                c2.number_format = '$#,##0.00'

    def _build_table_sheet(self, ws, parquet_path: Path):
        ws.views.sheetView[0].showGridLines = True
        cursor = self._conn.execute(f"SELECT * FROM read_parquet('{_safe(parquet_path)}')")
        cols = [d[0] for d in cursor.description]
        data = cursor.fetchall()

        header_fill = PatternFill("solid", fgColor="3B82F6")
        header_font = Font(bold=True, color="FFFFFF")

        for c_idx, col_name in enumerate(cols, 1):
            cell = ws.cell(row=1, column=c_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            ws.column_dimensions[openpyxl.utils.get_column_letter(c_idx)].width = max(len(col_name) + 5, 18)

        for r_idx, row in enumerate(data, 2):
            for c_idx, val in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                if isinstance(val, float):
                    cell.number_format = '$#,##0.00'
